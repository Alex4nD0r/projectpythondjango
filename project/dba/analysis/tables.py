import openpyxl


def get_table_data(file_name):
    workbook = openpyxl.load_workbook(file_name)
    worksheet = workbook.active
    data = []
    rows = worksheet.max_row
    cols = worksheet.max_column

    for i in range(0, rows):
        test = []
        for col in worksheet.iter_cols(1, worksheet.max_column):
            if col[i].value is None:
                test.append('')
            else:
                test.append(col[i].value)
        data.append(test)

    return data, rows, cols


